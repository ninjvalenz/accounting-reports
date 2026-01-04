"""
Highly Optimized Excel Parser v2 with:
- SQLite WAL mode for concurrent writes
- Pre-loaded caches (no race conditions)
- Connection per thread (thread-safe)
- Chunked batch inserts (handles huge files)
- Single file load with parallel processing
- Memory-efficient generators
"""

import os
import gc
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from database import get_connection, create_upload_record, update_upload_success, update_upload_error

# Thread-local storage for database connections
_thread_local = threading.local()

# Month mappings
MONTH_MAP = {
    "Jan'24": ("January", 2024), "Feb'24": ("February", 2024), "Mar'24": ("March", 2024),
    "Apr'24": ("April", 2024), "May'24": ("May", 2024), "Jun'24": ("June", 2024),
    "Jul'24": ("July", 2024), "Aug'24": ("August", 2024), "Sep'24": ("September", 2024),
    "Oct'24": ("October", 2024), "Nov'24": ("November", 2024), "Dec'24": ("December", 2024),
    "Jan'25": ("January", 2025), "Feb'25": ("February", 2025), "Mar'25": ("March", 2025),
    "Apr'25": ("April", 2025), "May'25": ("May", 2025), "Jun'25": ("June", 2025),
    "Jul'25": ("July", 2025), "Aug'25": ("August", 2025), "Sep'25": ("September", 2025),
    "Oct'25": ("October", 2025), "Nov'25": ("November", 2025), "Dec'25": ("December", 2025),
}

SHORT_MONTH_MAP = {
    'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
    'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
    'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
}

MONTH_ORDER = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# Global caches (pre-loaded before parallel processing)
_year_cache = {}
_month_cache = {}
_category_cache = {}
_product_cache = {}
_cache_lock = threading.Lock()

BATCH_SIZE = 500  # Insert in chunks


def get_thread_connection():
    """Get thread-local database connection with WAL mode"""
    if not hasattr(_thread_local, 'conn') or _thread_local.conn is None:
        conn = get_connection()
        conn.execute('PRAGMA journal_mode=WAL')  # Better concurrent performance
        conn.execute('PRAGMA synchronous=NORMAL')  # Faster writes
        conn.execute('PRAGMA cache_size=10000')  # Larger cache
        _thread_local.conn = conn
    return _thread_local.conn


def close_thread_connection():
    """Close thread-local connection"""
    if hasattr(_thread_local, 'conn') and _thread_local.conn:
        _thread_local.conn.close()
        _thread_local.conn = None


def reset_caches():
    """Reset all caches"""
    global _year_cache, _month_cache, _category_cache, _product_cache
    with _cache_lock:
        _year_cache = {}
        _month_cache = {}
        _category_cache = {}
        _product_cache = {}


def safe_float(val):
    """Convert to float, return 0 if invalid"""
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0


def safe_str(val, default=''):
    """Convert to string safely"""
    if val is None:
        return default
    return str(val).strip() or default


def parse_month(month_str):
    """Convert Excel month format to (month_name, year)"""
    return MONTH_MAP.get(month_str, (None, None))


def batch_insert(cursor, sql, data, batch_size=BATCH_SIZE):
    """Insert data in batches for better memory handling"""
    for i in range(0, len(data), batch_size):
        batch = data[i:i + batch_size]
        cursor.executemany(sql, batch)


def preload_caches():
    """Pre-load all reference data into caches"""
    global _year_cache, _month_cache, _category_cache, _product_cache
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Load months (static data)
    cursor.execute('SELECT id, name FROM months')
    for row in cursor.fetchall():
        _month_cache[row[1]] = row[0]
    
    # Load existing years
    cursor.execute('SELECT id, year FROM years')
    for row in cursor.fetchall():
        _year_cache[row[1]] = row[0]
    
    # Load existing categories
    cursor.execute('SELECT id, name FROM product_categories')
    for row in cursor.fetchall():
        _category_cache[row[1]] = row[0]
    
    # Load existing products
    cursor.execute('SELECT id, name, category_id FROM products')
    for row in cursor.fetchall():
        _product_cache[(row[1], row[2])] = row[0]
    
    conn.close()
    print(f"   📦 Cached: {len(_month_cache)} months, {len(_year_cache)} years, {len(_category_cache)} categories, {len(_product_cache)} products")


def get_or_create_year(cursor, year):
    """Get year ID with thread-safe caching"""
    if year in _year_cache:
        return _year_cache[year]
    
    with _cache_lock:
        # Double-check after acquiring lock
        if year in _year_cache:
            return _year_cache[year]
        
        cursor.execute('SELECT id FROM years WHERE year = ?', (year,))
        result = cursor.fetchone()
        if result:
            _year_cache[year] = result[0]
            return result[0]
        
        cursor.execute('INSERT INTO years (year) VALUES (?)', (year,))
        cursor.connection.commit()
        _year_cache[year] = cursor.lastrowid
        return cursor.lastrowid


def get_month_id(month_name):
    """Get month ID from cache (months are static)"""
    return _month_cache.get(month_name)


def get_or_create_category(cursor, category_name):
    """Get category ID with thread-safe caching"""
    if not category_name:
        category_name = "Uncategorized"
    
    if category_name in _category_cache:
        return _category_cache[category_name]
    
    with _cache_lock:
        if category_name in _category_cache:
            return _category_cache[category_name]
        
        cursor.execute('SELECT id FROM product_categories WHERE name = ?', (category_name,))
        result = cursor.fetchone()
        if result:
            _category_cache[category_name] = result[0]
            return result[0]
        
        cursor.execute('INSERT INTO product_categories (name) VALUES (?)', (category_name,))
        cursor.connection.commit()
        _category_cache[category_name] = cursor.lastrowid
        return cursor.lastrowid


def get_or_create_product(cursor, product_name, category_id, sub_category=None, type_of_sales=None):
    """Get product ID with thread-safe caching"""
    if not product_name:
        return None
    
    cache_key = (product_name, category_id)
    if cache_key in _product_cache:
        return _product_cache[cache_key]
    
    with _cache_lock:
        if cache_key in _product_cache:
            return _product_cache[cache_key]
        
        cursor.execute('SELECT id FROM products WHERE name = ? AND category_id = ?', (product_name, category_id))
        result = cursor.fetchone()
        if result:
            _product_cache[cache_key] = result[0]
            return result[0]
        
        cursor.execute('INSERT INTO products (name, category_id, sub_category, type_of_sales) VALUES (?, ?, ?, ?)',
                       (product_name, category_id, sub_category, type_of_sales))
        cursor.connection.commit()
        _product_cache[cache_key] = cursor.lastrowid
        return cursor.lastrowid


def read_all_sheets(filepath):
    """Read all sheets data at once (single file open)"""
    print("📖 Reading all sheets from Excel file...")
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheets_data = {}
    
    # Define sheets and their header rows
    sheet_configs = {
        'Day (in Month)': {'header_row': 1},
        'Sales Projection 2025': {'header_row': 2},
        'Data': {'header_row': 1},
        'Production Data': {'header_row': 1},
        'SALES BY FPR': {'header_row': 1},
    }
    
    for sheet_name, config in sheet_configs.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = config['header_row']
            
            # Get headers
            headers = []
            for row in ws.iter_rows(min_row=header_row, max_row=header_row):
                headers = [cell.value for cell in row]
                break
            
            # Read all rows into memory
            rows = []
            for row in ws.iter_rows(min_row=header_row + 1):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_data[headers[i]] = cell.value
                rows.append(row_data)
            
            sheets_data[sheet_name] = {'headers': headers, 'rows': rows}
            print(f"   ✅ {sheet_name}: {len(rows)} rows")
    
    # Read Dashboard-1 cost data separately (specific rows)
    if 'Dashboard-1' in wb.sheetnames:
        ws = wb['Dashboard-1']
        cost_rows = []
        for row in ws.iter_rows(min_row=110, max_row=120, max_col=3):
            cost_rows.append([cell.value for cell in row])
        sheets_data['Dashboard-1'] = {'cost_rows': cost_rows}
        print(f"   ✅ Dashboard-1: {len(cost_rows)} cost rows")
    
    wb.close()
    return sheets_data


def process_working_days(rows, upload_id):
    """Process working days data"""
    print("📅 Processing Working Days...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    batch = []
    for row in rows:
        month_name, year = parse_month(row.get('Months'))
        if not month_name:
            continue
        
        year_id = get_or_create_year(cursor, year)
        month_id = get_month_id(month_name)
        if month_id:
            batch.append((upload_id, year_id, month_id, int(safe_float(row.get('Days in months', 0)))))
            months_set.add(f"{month_name} {year}")
    
    batch_insert(cursor, 'INSERT OR REPLACE INTO working_days (upload_id, year_id, month_id, days) VALUES (?, ?, ?, ?)', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    return months_set


def process_sales_projection(rows, headers, upload_id):
    """Process sales projection data"""
    print("📊 Processing Sales Projection...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    # Find month columns
    month_columns = [h for h in headers if h and "'25" in str(h) and "." not in str(h)]
    
    batch = []
    for row in rows:
        product_name = safe_str(row.get('Products'))
        if not product_name:
            continue
        
        category_id = get_or_create_category(cursor, safe_str(row.get('Product Category')))
        product_id = get_or_create_product(cursor, product_name, category_id,
                                           safe_str(row.get('Product Category 2')) or None,
                                           safe_str(row.get('Type of Sales')) or None)
        if not product_id:
            continue
        
        for month_col in month_columns:
            month_name, year = parse_month(month_col)
            if not month_name:
                continue
            year_id = get_or_create_year(cursor, year)
            month_id = get_month_id(month_name)
            if month_id:
                batch.append((upload_id, year_id, month_id, product_id, safe_float(row.get(month_col, 0))))
                months_set.add(f"{month_name} {year}")
    
    batch_insert(cursor, 'INSERT OR REPLACE INTO budget_projection (upload_id, year_id, month_id, product_id, quantity) VALUES (?, ?, ?, ?, ?)', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    return months_set


def process_sales_data(rows, upload_id):
    """Process sales data with aggregation"""
    print("💰 Processing Sales Data...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    # Aggregate in memory
    aggregated = {}
    for row in rows:
        month_name, year = parse_month(row.get('Month'))
        if not month_name:
            continue
        
        product_name = safe_str(row.get('Products'))
        if not product_name:
            continue
        
        year_id = get_or_create_year(cursor, year)
        month_id = get_month_id(month_name)
        category_id = get_or_create_category(cursor, safe_str(row.get('Product Category')))
        product_id = get_or_create_product(cursor, product_name, category_id,
                                           safe_str(row.get('Product Category 2')) or None,
                                           safe_str(row.get('Type of Sales')) or None)
        
        if not month_id or not product_id:
            continue
        
        key = (year_id, month_id, product_id)
        if key not in aggregated:
            aggregated[key] = [0.0] * 6
        
        aggregated[key][0] += safe_float(row.get('Qty-Budget', 0))
        aggregated[key][1] += safe_float(row.get('Amount-Budget (US$)', 0))
        aggregated[key][2] += safe_float(row.get('Qty-Actual', 0))
        aggregated[key][3] += safe_float(row.get('Amount-Actual (US$)', 0))
        aggregated[key][4] += safe_float(row.get('Qty in Liters (Budgeted)', 0))
        aggregated[key][5] += safe_float(row.get('Qty in Liters', 0))
        months_set.add(f"{month_name} {year}")
    
    # Batch insert
    batch = [(upload_id, k[0], k[1], k[2], v[0], v[1], v[2], v[3], v[4], v[5]) for k, v in aggregated.items()]
    batch_insert(cursor, '''INSERT INTO sales_data (upload_id, year_id, month_id, product_id, qty_budget, 
                  amount_budget, qty_actual, amount_actual, qty_liters_budget, qty_liters_actual) 
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    del aggregated
    return months_set


def process_production_data(rows, upload_id):
    """Process production data with aggregation"""
    print("🏭 Processing Production Data...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    aggregated = {}
    for row in rows:
        month_name, year = parse_month(row.get('Month'))
        if not month_name:
            continue
        
        product_name = safe_str(row.get('Products'))
        if not product_name:
            continue
        
        year_id = get_or_create_year(cursor, year)
        month_id = get_month_id(month_name)
        category_id = get_or_create_category(cursor, safe_str(row.get('Product Category')))
        product_id = get_or_create_product(cursor, product_name, category_id,
                                           safe_str(row.get('Product Category 2')) or None,
                                           safe_str(row.get('Type of Sales')) or None)
        
        if not month_id or not product_id:
            continue
        
        key = (year_id, month_id, product_id)
        if key not in aggregated:
            aggregated[key] = [0.0] * 4
        
        aggregated[key][0] += safe_float(row.get('Qty-Budgeted', 0))
        aggregated[key][1] += safe_float(row.get('Qty Budgeted (In Ltrs)', 0))
        aggregated[key][2] += safe_float(row.get('Qty-Actual', 0))
        aggregated[key][3] += safe_float(row.get('Qty in Liters', 0))
        months_set.add(f"{month_name} {year}")
    
    batch = [(upload_id, k[0], k[1], k[2], v[0], v[1], v[2], v[3]) for k, v in aggregated.items()]
    batch_insert(cursor, '''INSERT INTO production_data (upload_id, year_id, month_id, product_id, qty_budget, 
                  qty_budget_liters, qty_actual, qty_actual_liters) VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    del aggregated
    return months_set


def process_sales_by_fpr(rows, upload_id):
    """Process sales by FPR data"""
    print("👥 Processing Sales by FPR...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    batch = []
    for row in rows:
        year = row.get('Year')
        month_short = row.get('Month')
        if year is None or month_short is None:
            continue
        
        month_name = SHORT_MONTH_MAP.get(month_short)
        if not month_name:
            continue
        
        year_id = get_or_create_year(cursor, int(year))
        month_id = get_month_id(month_name)
        if not month_id:
            continue
        
        batch.append((
            upload_id, year_id, month_id,
            safe_str(row.get('SalesMan'), 'Unknown'),
            safe_str(row.get('Location'), 'Unknown'),
            safe_str(row.get('Type of sales'), 'Unknown'),
            safe_float(row.get('Amount', 0))
        ))
        months_set.add(f"{month_name} {int(year)}")
    
    batch_insert(cursor, '''INSERT INTO sales_by_fpr (upload_id, year_id, month_id, salesman, location, 
                  type_of_sales, amount) VALUES (?, ?, ?, ?, ?, ?, ?)''', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    return months_set


def process_cost_data(cost_rows, upload_id):
    """Process cost data"""
    print("💰 Processing Cost Data...")
    conn = get_thread_connection()
    cursor = conn.cursor()
    months_set = set()
    
    batch = []
    for row in cost_rows:
        month_cell = row[0]
        if not month_cell or not isinstance(month_cell, str):
            continue
        
        month_name, year = parse_month(month_cell)
        if not month_name:
            continue
        
        year_id = get_or_create_year(cursor, year)
        month_id = get_month_id(month_name)
        if month_id:
            batch.append((upload_id, year_id, month_id, safe_float(row[1]), safe_float(row[2])))
            months_set.add(f"{month_name} {year}")
    
    batch_insert(cursor, 'INSERT OR REPLACE INTO cost_data (upload_id, year_id, month_id, fuel, lec) VALUES (?, ?, ?, ?, ?)', batch)
    conn.commit()
    
    print(f"   ✅ Saved {len(batch)} records")
    return months_set


def process_excel_file(filepath):
    """Main function - optimized parallel processing"""
    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"🚀 OPTIMIZED EXCEL PROCESSOR v2")
    print(f"📁 File: {filename}")
    print(f"{'='*60}")
    
    # Reset and preload caches
    reset_caches()
    preload_caches()
    
    upload_id = create_upload_record(filename)
    sheets_processed = []
    all_months = set()
    
    try:
        # Phase 1: Read all sheets at once (single file open)
        sheets_data = read_all_sheets(filepath)
        gc.collect()
        
        # Phase 2: Process reference data first (sequential - builds cache)
        print("\n📋 Phase 1: Processing reference data...")
        
        if 'Day (in Month)' in sheets_data:
            months = process_working_days(sheets_data['Day (in Month)']['rows'], upload_id)
            all_months.update(months)
            sheets_processed.append("Day (in Month)")
        
        if 'Sales Projection 2025' in sheets_data:
            data = sheets_data['Sales Projection 2025']
            months = process_sales_projection(data['rows'], data['headers'], upload_id)
            all_months.update(months)
            sheets_processed.append("Sales Projection 2025")
        
        gc.collect()
        
        # Phase 3: Process data sheets in parallel
        print("\n🚀 Phase 2: Processing data sheets in parallel...")
        
        def process_wrapper(func, data, upload_id, name):
            """Wrapper to handle thread-local connections"""
            try:
                result = func(data, upload_id)
                return name, result, None
            except Exception as e:
                return name, set(), str(e)
            finally:
                close_thread_connection()
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = []
            
            if 'Data' in sheets_data:
                futures.append(executor.submit(process_wrapper, process_sales_data, 
                                               sheets_data['Data']['rows'], upload_id, "Data"))
            
            if 'Production Data' in sheets_data:
                futures.append(executor.submit(process_wrapper, process_production_data,
                                               sheets_data['Production Data']['rows'], upload_id, "Production Data"))
            
            if 'SALES BY FPR' in sheets_data:
                futures.append(executor.submit(process_wrapper, process_sales_by_fpr,
                                               sheets_data['SALES BY FPR']['rows'], upload_id, "SALES BY FPR"))
            
            for future in as_completed(futures):
                name, months, error = future.result()
                if error:
                    print(f"   ❌ {name}: {error}")
                else:
                    all_months.update(months)
                    sheets_processed.append(name)
        
        gc.collect()
        
        # Phase 4: Process cost data
        print("\n📋 Phase 3: Processing cost data...")
        if 'Dashboard-1' in sheets_data:
            months = process_cost_data(sheets_data['Dashboard-1']['cost_rows'], upload_id)
            all_months.update(months)
            sheets_processed.append("Dashboard-1")
        
        # Sort months/years
        months_years_list = sorted(list(all_months),
                                   key=lambda x: (int(x.split()[-1]), MONTH_ORDER.index(x.split()[0])))
        
        update_upload_success(upload_id, sheets_processed, months_years_list)
        
        print(f"\n{'='*60}")
        print(f"✅ COMPLETE!")
        print(f"   📊 Sheets: {len(sheets_processed)}")
        print(f"   📅 Periods: {len(months_years_list)}")
        print(f"   🗃️  Cached: {len(_category_cache)} categories, {len(_product_cache)} products")
        print(f"{'='*60}\n")
        
        return {
            'success': True,
            'upload_id': upload_id,
            'sheets_processed': sheets_processed,
            'months_years_processed': months_years_list
        }
        
    except Exception as e:
        update_upload_error(upload_id, str(e))
        print(f"\n❌ Error: {str(e)}")
        raise e
    finally:
        reset_caches()
        close_thread_connection()
        gc.collect()


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        process_excel_file(sys.argv[1])
    else:
        print("Usage: python excel_parser.py <path_to_excel_file>")
